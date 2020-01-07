import subprocess
import sys
import os
from openpyxl import Workbook
from openpyxl import load_workbook




def parsing_res(site,WS):
	global Row_in_excel_file
	try:
		f = open('res_nmap.txt', 'r')
	except:
		print("No subdomains found!")
		sys.exit(0)
	os.system('cls' if os.name == 'nt' else 'clear')
	WS.cell(Row_in_excel_file,1).value = site
	lines = []
	lines = f.read().splitlines()
	f.close()
	IP_start = lines[1].find('(')
	IP_end = lines[1].find(')')
	IP = lines[1][IP_start+1:IP_end]
	WS.cell(Row_in_excel_file,2).value = IP
	WS.cell(Row_in_excel_file,3).value = ''
	for i in range(6,len(lines)):
		if lines[i].find("/tcp") != -1 or lines[i].find("/udp") != -1:
			first_space = lines[i].find('/')
			PORT= lines[i][:first_space]
			WS.cell(Row_in_excel_file,3).value += PORT
			first_space = lines[i].find(' ')
			while lines[i][first_space] == ' ':
				first_space += 1
			second_space = lines[i].find(' ',first_space)
			WS.cell(Row_in_excel_file,3).value += ' '
			STATE = lines[i][first_space:second_space]
			

			while lines[i][second_space] == ' ':
				second_space += 1
			first_space = lines[i].find(' ',second_space)

			if first_space == -1:
				SERVICE = lines[i][second_space:]
				WS.cell(Row_in_excel_file,3).value += SERVICE
				VERSION = '-'
				WS.cell(Row_in_excel_file,3).value += '\n'
			else:
				SERVICE = lines[i][second_space:first_space]
				WS.cell(Row_in_excel_file,3).value += SERVICE
				WS.cell(Row_in_excel_file,3).value += ' '
				while lines[i][first_space] == ' ':
					first_space += 1
				VERSION = lines[i][first_space:]
				WS.cell(Row_in_excel_file,3).value += VERSION
				WS.cell(Row_in_excel_file,3).value += '\n'
		elif lines[i].find("CVE-") != -1:
			CVE_S = lines[i].find("CVE-")
			CVE_E = lines[i].find('\t', CVE_S)
			CVE = lines[i][CVE_S:CVE_E]
			WS.cell(Row_in_excel_file,3).value += '\t'
			WS.cell(Row_in_excel_file,3).value += CVE
			WS.cell(Row_in_excel_file,3).value += '\n'
	Row_in_excel_file += 1

def full_anlyz( site,WS):
	if SYSTEM_Win:
		out = 'nmap -sV --script vulners.nse -oN res_nmap.txt '+ site
		res = subprocess.call(out,shell = True)
		os.system('cls' if os.name == 'nt' else 'clear')
		print("Result receved. Preparing for saving information")
		parsing_res(site,WS)
		print("Information added in table")
	else:
		out = 'nmap -sV --script vulners.nse -oN res_nmap.txt '+ site
		res = subprocess.call(out,shell = True)
		os.system('cls' if os.name == 'nt' else 'clear')
		print("Result receved. Preparing for saving information")
		parsing_res(site,WS)
		print("Information added in table")

			


def check_domain():
	os.system('cls' if os.name == 'nt' else 'clear')
	print("Please, enter domain you want to check")
	print("Example: google.com")
	print("> ",end ='')
	site = input()
	os.system('cls' if os.name == 'nt' else 'clear')
	print("You entered: " + site, end = '\n\n')
	print("Do you want to find all subdomains or only resolvable subdomains?")
	print("Eneter \"All\" - to find all subdomains and \"Res\" - for resolvable subdomains")
	print("> ", end ='')
	mode = input()
	if mode.lower() == "res":
		if SYSTEM_Win:
			out = 'findomain -t ' + site + ' -r -u sub_res.txt'
			os.system('cls' if os.name == 'nt' else 'clear')
			subprocess.call(out)
			
		else:
			out = './findomain -t ' + site + ' -r -u sub_res.txt'
			os.system('cls' if os.name == 'nt' else 'clear')
			subprocess.call(out,shell =True)
			
	else:
		if SYSTEM_Win:
			out = 'findomain -t ' + site + ' -u sub_res.txt'
			os.system('cls' if os.name == 'nt' else 'clear')
			subprocess.call(out)
			
		else:
			out = './findomain -t ' + site + ' -u sub_res.txt'
			os.system('cls' if os.name == 'nt' else 'clear')
			subprocess.call(out,shell =True)
			
	os.system('cls' if os.name == 'nt' else 'clear')
	input("Press Enter to continue...")
	os.system('cls' if os.name == 'nt' else 'clear')

	try:
		f = open('sub_res.txt', 'r')
	except:
		print("No subdomains found!")
		return

	lines = []
	lines = f.read().splitlines()
	f.close()
	print("Found subdomains:",end = '\n\n')
	for i in range(len(lines)):
		print(lines[i])
	print()
	input("Press Enter to continue...")
	os.system('cls' if os.name == 'nt' else 'clear')
	if mode.lower() == all:
		return
	print("Please, input name of excel file for save.")
	print("Example: test.xlsx")
	excel_name = input("> ")
	os.system('cls' if os.name == 'nt' else 'clear')
	ExcelBook = Workbook()
	WorkSheet = ExcelBook.active
	
	print("Starting analaize subdomains...")
	for i in range(len(lines)):
		print("Analyzing: "+lines[i]+"...")
		full_anlyz(lines[i],WorkSheet)
	ExcelBook.save(excel_name)
	print("All information saved in " + excel_name)


	
def add_token():
	os.system('cls' if os.name == 'nt' else 'clear')
	

	print("Please, enter what type of token you want to add:")
	print("FB - for Facebook API\nSpyse - for Spyse API\nVirustotal - for Virustotal API")
	type = input()
	if type.lower() == "fb":
		print("Please, enter your token:", end = ' ')
		token = input()
		if SYSTEM_Win:
			token.replace("|","^|")
			out = 'set findomain_fb_token='+token
			res = os.system(out)
		else:
			out = 'export findomain_fb_token="'+token+'"'
			res = os.system(out,shell =True)

	elif type.lower() == "spyse":
		print("Please, enter your token:", end = ' ')
		token = input()
		if SYSTEM_Win:
			token.replace("|","^|")
			out = 'set findomain_spyse_token='+token
			res = os.system(out)
		else:
			out = 'export findomain_spyse_token="'+token+'"'
			res = os.system(out,shell =True)

	else:
		print("Please, enter your token:", end = ' ')
		token = input()
		if SYSTEM_Win:
			token.replace("|","^|")
			out = 'set findomain_virustotal_token='+token
			res = os.system(out)
			
		else:
			out = 'export findomain_virustotal_token="'+token+'"'
			res = os.system(out,shell =True)
			
	if res == 0:
		print("Token successfully added!")
	else:
		print("Problems with adding token")


Row_in_excel_file = 1

SYSTEM_Win = True
if sys.platform == "linux" or sys.platform == "linux2":
    SYSTEM_Win= False
elif sys.platform == "darwin":
    SYSTEM_Win= False

while 1:

	os.system('cls' if os.name == 'nt' else 'clear')
	if SYSTEM_Win:
		print("U have Win Sys",end = '\n\n')
	else:
		print("U have smthng else",end = '\n\n')
	print("Hi there! Please enter one of the commands:")
	print("check - for analyze domain")
	print("token - for adding one of tokens")
	print("exit - for out of plugin")
	print("> ",end='')
	command = input()

	if command.lower() == "check": 
		check_domain()
	elif command.lower() == "exit":
		sys.exit(0)
	elif command.lower() == "token":
		add_token()
	else:
		os.system('cls' if os.name == 'nt' else 'clear')
		print("Unknown command: " + command +" please try again")

	input("Press Enter to continue...")